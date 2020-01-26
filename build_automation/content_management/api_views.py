import os
from openpyxl import Workbook
from io import BytesIO

from django.core.files.base import ContentFile
from django.http import Http404, FileResponse

from rest_framework import status
from rest_framework.mixins import CreateModelMixin, ListModelMixin, RetrieveModelMixin
from rest_framework.response import Response
from rest_framework.reverse import reverse
from rest_framework.viewsets import ModelViewSet, ViewSet

from content_management.exceptions import DuplicateContentFileException
from content_management.models import (
    Build, Cataloger, Content, Coverage, Creator, Directory, DirectoryLayout,
    Keyword, Language, Subject, Workarea, MetadataSheet, Collection
)
from content_management.serializers import (
    BuildSerializer, CatalogerSerializer, ContentSerializer, CoverageSerializer, CreatorSerializer,
    DirectoryLayoutSerializer, DirectorySerializer, KeywordSerializer, LanguageSerializer, SubjectSerializer,
    WorkareaSerializer, MetadataSheetSerializer, CollectionSerializer
)
from content_management.filters import ContentsFilterSet
from content_management.tasks import start_dirlayout_build
from content_management.utils import DiskSpace, LibraryVersionBuildUtil, temporary_filename


class ContentApiViewset(ModelViewSet):
    queryset = Content.objects.all()
    serializer_class = ContentSerializer
    search_fields = ('name', 'description')
    filterset_class = ContentsFilterSet

    def get_queryset(self, request=None):
        qs = super().get_queryset()
        if request is None:
            self.filterset = self.filterset_class(self.request.GET, queryset=qs)
        else:
            self.filterset = self.filterset_class(request.GET, queryset=qs)
        return self.filterset.qs.distinct()

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except DuplicateContentFileException as dup:
            content_url = reverse('content-detail', args=[dup.content.pk], request=request)
            data = {
                'result': 'error',
                'error': 'DUPLICATE_FILE_UPLOADED',
                'existing_content': {
                    'content_url': content_url,
                    'file_url': request.build_absolute_uri(dup.content.content_file.url)
                }
            }
            return Response(data, status=status.HTTP_409_CONFLICT)

    def update(self, request, *args, **kwargs):
        try:
            return super().update(request, *args, **kwargs)
        except DuplicateContentFileException as dup:
            content_url = reverse('content-detail', args=[dup.content.pk], request=request)
            data = {
                'result': 'error',
                'error': 'DUPLICATE_FILE_UPLOADED',
                'existing_content': {
                    'content_url': content_url,
                    'file_url': request.build_absolute_uri(dup.content.content_file.url)
                }
            }
            return Response(data, status=status.HTTP_409_CONFLICT)


class CreatorViewSet(ModelViewSet):
    serializer_class = CreatorSerializer
    queryset = Creator.objects.all()


class CoverageViewSet(ModelViewSet):
    serializer_class = CoverageSerializer
    queryset = Coverage.objects.all()


class CollectionViewSet(ModelViewSet):
    serializer_class = CollectionSerializer
    queryset = Collection.objects.all()


class SubjectViewSet(ModelViewSet):
    serializer_class = SubjectSerializer
    queryset = Subject.objects.all()


class KeywordViewSet(ModelViewSet):
    serializer_class = KeywordSerializer
    queryset = Keyword.objects.all()


class WorkareaViewSet(ModelViewSet):
    serializer_class = WorkareaSerializer
    queryset = Workarea.objects.all()


class LanguageViewSet(ModelViewSet):
    serializer_class = LanguageSerializer
    queryset = Language.objects.all()


class CatalogerViewSet(ModelViewSet):
    serializer_class = CatalogerSerializer
    queryset = Cataloger.objects.all()


class DirectoryLayoutViewSet(ModelViewSet):
    serializer_class = DirectoryLayoutSerializer
    queryset = DirectoryLayout.objects.all()


class DirectoryViewSet(ModelViewSet):
    serializer_class = DirectorySerializer
    queryset = Directory.objects.all()


class DirectoryCloneApiViewSet(ViewSet, CreateModelMixin):
    serializer_class = DirectoryLayoutSerializer
    CLONE_SUFFIX = "_clone"

    def create(self, request, *args, **kwargs):
        original_layout = self.get_queryset()
        if(DirectoryLayout.objects.filter(name=original_layout.name + self.CLONE_SUFFIX).count() >= 1):
            dir = DirectoryLayout.objects.get(name=original_layout.name + self.CLONE_SUFFIX)
            layout_url = reverse('directorylayout-detail', args=[dir.id], request=request)
            data = {
                'result': 'error',
                'error': 'DIRECTORY_LAYOUT_ALREADY_EXISTS',
                'existing_directory_layout': {
                    'directory_layout_name': original_layout.name + self.CLONE_SUFFIX,
                    'directory_layout': layout_url
                }
            }
            return Response(data, status=status.HTTP_409_CONFLICT)
        cloned_layout = DirectoryLayout(
            name=original_layout.name + self.CLONE_SUFFIX, description=original_layout.description
        )
        cloned_layout.pk = None
        dup_banner = ContentFile(original_layout.banner_file.read())
        dup_banner.name = original_layout.original_file_name
        cloned_layout.banner_file = dup_banner
        cloned_layout.save()

        dir_queryset = Directory.objects.filter(dir_layout=original_layout, parent=None)
        self.__clone_directory_tree(None, cloned_layout, dir_queryset, None)

        serializer = DirectoryLayoutSerializer(cloned_layout, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def get_queryset(self, **kwargs):
        queryset = DirectoryLayout.objects.get(id=self.kwargs['id'])
        return queryset

    def __clone_directory_tree(
            self, filter_criteria_util, cloned_dir_layout,
            original_directories, parent_cloned_directory):
        for each_original_directory in original_directories:
            cloned_directory = Directory(name=each_original_directory.name)
            cloned_directory.dir_layout = cloned_dir_layout
            cloned_directory.parent = parent_cloned_directory
            if (
                each_original_directory.banner_file is not None and
                len(each_original_directory.banner_file.name) > 0 and
                os.path.exists(each_original_directory.banner_file.path)
            ):
                dup_banner = ContentFile(each_original_directory.banner_file.read())
                dup_banner.name = each_original_directory.original_file_name
                cloned_directory.banner_file = dup_banner
            cloned_directory.save()
            cloned_directory.individual_files.set(list(each_original_directory.individual_files.all()))
            cloned_directory.creators.set(list(each_original_directory.creators.all()))
            cloned_directory.coverages.set(list(each_original_directory.coverages.all()))
            cloned_directory.subjects.set(list(each_original_directory.subjects.all()))
            cloned_directory.keywords.set(list(each_original_directory.keywords.all()))
            cloned_directory.workareas.set(list(each_original_directory.workareas.all()))
            cloned_directory.languages.set(list(each_original_directory.languages.all()))
            cloned_directory.catalogers.set(list(each_original_directory.catalogers.all()))
            cloned_directory.save()
            self.__clone_directory_tree(
                filter_criteria_util, cloned_dir_layout,
                each_original_directory.subdirectories.all(), cloned_directory
            )


class AllTagsApiViewSet(ViewSet, ListModelMixin):
    """
    Get all kinds of tags in a single API call
    creator, coverage, subjects, workareas, keywords, language and cataloger
    """
    def list(self, request, *args, **kwarsgs):
        response_data = {
            'creators': Creator.objects.all().values(),
            'coverages': Coverage.objects.all().values(),
            'subjects': Subject.objects.all().values(),
            'keywords': Keyword.objects.all().values(),
            'workareas': Workarea.objects.all().values(),
            'languages': Language.objects.all().values(),
            'catalogers': Cataloger.objects.all().values(),
            'collections': Collection.objects.all().values(),
        }
        return Response(response_data, status=status.HTTP_200_OK)


class BuildLibraryVersionViewSet(ViewSet, CreateModelMixin, RetrieveModelMixin):
    """
    Start the build process on the
    """

    def create(self, request, *args, **kwargs):
        build_util = LibraryVersionBuildUtil()
        latest_build = build_util.get_latest_build()
        if latest_build is None or latest_build.task_state == Build.TaskState.FINISHED:
            dir_layout_id = self.kwargs['id']
            start_dirlayout_build.delay(dir_layout_id)
            return Response(
                {
                    'status': 'successful',
                    'message': 'Build process has been successfully started'
                }
            )
        else:
            return Response(
                {
                    'status': 'failure',
                    'message': 'Another build is already in progress. Please wait until it completes'
                },
                status=status.HTTP_409_CONFLICT
            )

    def list(self, request, *args, **kwargs):
        build_util = LibraryVersionBuildUtil()
        latest_build = build_util.get_latest_build()
        if latest_build is None:
            return Response([])
        serializer = BuildSerializer(latest_build, context={'request': request})
        return Response([
            serializer.data
        ])


class DiskSpaceViewSet(ViewSet):
    def list(self, request):
        dp = DiskSpace()
        data = {
            'available_space': dp.getfreespace()[0],
            'total_space': dp.getfreespace()[1]
        }
        return Response(data)


class MetadataSheetApiViewSet(ModelViewSet):
    serializer_class = MetadataSheetSerializer
    queryset = MetadataSheet.objects.all()

    def create(self, request, *args, **kwargs):
        try:
            return super().create(request, *args, **kwargs)
        except DuplicateContentFileException as dup:
            metadata_url = reverse('metadata', args=[dup.content.pk], request=request)
            data = {
                'result': 'error',
                'error': 'DUPLICATE_FILE_UPLOADED',
                'existing_metadata': {
                    'metadata_url': metadata_url,
                    'file_url': request.build_absolute_uri(dup.content.metadata_file.url)
                }
            }
            return Response(data, status=status.HTTP_409_CONFLICT)


class MetadataMatchViewSet(ViewSet):
    queryset = Content.objects.values_list('name', flat=True)

    # print(list(queryset))
    def getNames(self, request):
        fileNameArray = list(self.queryset)
        data = {
            "content_files": fileNameArray
        }

        return Response(data)


class SpreadsheetView(ViewSet):
    # order integer is needed because order in a python dictionary is otherwise arbitrary
    details_map = {
        'contents': {
            'viewset_class': ContentApiViewset,
            'fields': {
                'name': ['field', 0],
                'description': ['field', 1],
                'updated_time': ['field', 2],
                'creators': ['many', 3],
                'coverage': ['foreign', 4],
                'subjects': ['many', 5],
                'keywords': ['many', 6],
                'workareas': ['many', 7],
                'language': ['foreign', 8],
                'cataloger': ['foreign', 9],
                'active': ['field', 10],
                'audience': ['field', 11],
                'collections': ['many', 12],
            }
        }
    }

    # Helper function that returns the string to be placed in the workbook
    # based on a Content instance, with a given field name and type
    def get_data_string(self, content, field_name, field_type):
        if field_type == "field":
            return getattr(content, field_name, "")
        elif field_type == "many":
            all_related = getattr(content, field_name).all()
            if len(all_related) > 0:
                return ", ".join([related.name for related in all_related])
            else:
                return ""
        elif field_type == "foreign":
            return getattr(getattr(content, field_name), 'name', "")

    # View function for django
    def getSpreadsheet(self, request, *args, **kwargs):
        sheet_type = str(kwargs['type'])
        try:
            details = self.details_map[sheet_type]
            fields = details["fields"]
            viewset = details["viewset_class"]().get_queryset(request)

            data_table = [
                [
                    str(self.get_data_string(content, field_name, field_type[0]))
                    for field_name, field_type in fields.items()
                ]
                for content in viewset
            ]

            filename = sheet_type + '.xlsx'
            return FileResponse(self.build_xlsx(data_table, fields), as_attachment=True, filename=filename)
        except LookupError:
            raise Http404()

    # Helper function that takes a data table and column fields and returns bytes for an xlsx file
    def build_xlsx(self, data_table, fields):
        with temporary_filename() as filename:
            wb = Workbook()
            ws = wb.active

            sorted_fields = sorted(fields.keys(), key=lambda e: fields[e][1])

            for i, field in enumerate(sorted_fields):
                _ = ws.cell(row=1, column=i+1, value=field)

            for i, row in enumerate(data_table):
                for k, field_value in enumerate(row):
                    _ = ws.cell(row=i+2, column=k+1, value=field_value)

            wb.save(filename)
            with open(filename, mode='rb') as f:
                return BytesIO(f.read())
